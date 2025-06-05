import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
from datetime import datetime
import xlwings as xw
import pandas as pd
import os
import constants as cs
import time

# value pricing is added to the dataframe during dataframe transformation.
# but we want to call it out visually in the xlsx form with a yellow highlight
def update_value_pricing_bg(sheet):
    
    vp_dict = cs.value_pricing
    
    yell_hex = 'FFFF65'
    yell_fill = PatternFill(start_color=yell_hex, end_color=yell_hex, fill_type='solid')

    for row in sheet.iter_rows():
        strain_value = row[2].value
        price_value = row[11].value
        cell_to_update = row[11]

        for key, val in vp_dict.items():      
            if strain_value == key[1] and price_value == val:
                cell_to_update.fill = yell_fill
    
    return sheet

# the downloaded report from acu doesn't play well with the code
# you could manually resave and the file is then fine
# this effectively does that 
def clean_excel_file(input_path, output_path):
    app = xw.App(visible=False)
    try:
        wb = app.books.open(input_path)
        wb.save(output_path)
        wb.close()
    finally:
        app.quit()
    return output_path


# merge like product description columns so the same description isn't listed over and over and over
def merge_cells_in_column(sheet, column_letter, start_row):

    column_index = openpyxl.utils.column_index_from_string(column_letter)
    
    current_value = sheet.cell(row=start_row, column=column_index).value
    
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        
        if cell_value == current_value:
            continue
        else:
            if row - start_row > 1:
              sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=row - 1, end_column=column_index)
            
            start_row = row
            current_value = cell_value
    if sheet.max_row - start_row >= 0:
        if sheet.max_row + 1 - start_row > 1:
            sheet.merge_cells(start_row=start_row, start_column=column_index, end_row=sheet.max_row, end_column=column_index)

    return sheet


# add grey background to empty fields that aren't already red
def grey_out_cells(sheet, last_total_row):
    
    red_hex = '00D60000'
    grey_hex = '545454'
    grey_fill = PatternFill(start_color=grey_hex, end_color=grey_hex, fill_type='solid')

    for row in sheet.iter_rows(min_row=8, max_row=last_total_row, min_col=4, max_col=11):
        for cell in row:
            if cell.value is None or cell.value == '' or cell.value == ' ':
                current_fill = cell.fill.start_color.rgb
                # print(current_fill)
                if current_fill != red_hex:
                    cell.fill = grey_fill

# first/inital formatting to get the spreadsheet formatting started - the dataframe will be added to this
def format_white_bg(sheet, cell_range):
    # formats a range of cells to have no borders and white fill
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    no_border = Border(
        left=Side(border_style=None),
        right=Side(border_style=None),
        top=Side(border_style=None),
        bottom=Side(border_style=None)
    )
    white_fill = PatternFill(fill_type='solid', start_color="FFFFFF", end_color="FFFFFF")

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = no_border
            cell.fill = white_fill

# the columns after initial importing of the dataframe into the xlsx are fixed and narrow
# this will get the maxlength of the strings in the column and widen the column for better visibility 
def adjust_column_width(sheet):

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            if isinstance(cell.value, float):
              cell_length = len(str(cell.value))
              max_length = max(max_length, cell_length)
            else:
              cell_length = len(str(cell.value))
              max_length = max(max_length, cell_length)

        adjusted_width = (max_length + 1)
        sheet.column_dimensions[column_letter].width = adjusted_width

def center_align_columns(sheet):
    for column in sheet.columns:
        for cell in column:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# this is leveraged in a few different functions
def get_max_total_row(sheet):
    last_row = 0
    
    for row in range(9, sheet.max_row + 1):
        sstr = '=IFERROR'
        cell_value = sheet.cell(row=row, column=18).value
        if sstr in str(cell_value):
            last_row = row    
    return last_row

def add_borders(sheet, last_total_row):
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    
    for row in sheet.iter_rows(min_row=8, max_row=last_total_row, min_col=2, max_col=18):
        for cell in row:
            # if cell.value is not None:
            cell.border = thin_border

# this creates the Order Total field that sums everything ordered
def add_total_sum(sheet, last_total_row):
    
    black_bold_font = Font(name='Calibri', size=12, bold=True, italic=False, color='000000')
    total_cell_row = last_total_row + 1
    total_sum_cell = sheet[f'R{total_cell_row}']
    total_sum_cell.value = f'=SUM(R9:R{last_total_row})'
    total_sum_cell.font = black_bold_font

    total_text_cell = sheet[f'Q{total_cell_row}']
    total_text_cell.value = 'ORDER TOTAL'
    total_text_cell.font = black_bold_font

# convert float values to %
# def convert_float_percentage(sheet):
#     for row in sheet:
#         for cell in row:
#             if isinstance(cell.value, float):
#                 cell.value = cell.value / 100
#                 cell.number_format = '0.0%'

def convert_float_percentage(sheet):
    header_row_idx = 7
    headers = [cell.value for cell in sheet[header_row_idx]]
    target_columns = {'THC-A', 'Total THC', 'Total Terpenes', 'TAC'}

    for row in sheet.iter_rows(min_row=header_row_idx + 1):  # start from row 8
        for cell in row:
            col_idx = cell.col_idx - 1  # 0-based index
            if col_idx < len(headers) and headers[col_idx] in target_columns:
                if isinstance(cell.value, float):
                    cell.value = cell.value / 100
                    cell.number_format = '0.0%'
                elif isinstance(cell.value, int):
                    cell.value = f'{cell.value}.0%'

# when insert_start_row in df_functions runs it creates several duplicate
# separator rows - this cleans that up - also cleans up purple and green dupe instances
def delete_dupe_red_rows(sheet):

    previous_value = None
    rows_to_delete = []

    for row in sheet.iter_rows(min_row=8):
        if row[1].value is None or row[1].value == "":
            if previous_value is not None and row[2].value == previous_value:
                rows_to_delete.append(row[0].row)
            else:
                previous_value = row[2].value

    # delete the marked rows - reverse order to avoid index shifting
    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)
    
    return sheet

# add grey bg to column headers
def grey_headers(sheet):
    
    cell_range = 'B7:R7'
    grey_hex = 'BABABA'
    grey_fill = PatternFill(start_color=grey_hex, end_color=grey_hex, fill_type='solid')
    
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = grey_fill

    return sheet

# poorly named - should be "color separator rows"
# cfx gummies gets green
# muze gets purple
# the rest red  
def add_separator_row(sheet):

    red_hex = 'D60000'
    purple_hex = '7030A0'
    green_hex = '00B050'
    stop_column = 18
    red_fill = PatternFill(start_color=red_hex, end_color=red_hex, fill_type='solid')
    green_fill = PatternFill(start_color=green_hex, end_color=green_hex, fill_type='solid')
    purple_fill = PatternFill(start_color=purple_hex, end_color=purple_hex, fill_type='solid')

    for row in sheet.iter_rows(min_row=8):
        first_cell = row[0].value
        second_cell = row[1].value
        third_cell = row[2].value

        # check conditions for color
        if first_cell is None and second_cell is None:
            if third_cell == "CuratedFX Gummies - Rapid Onset - 100mg THC" or third_cell == "CuratedFX Stir Stix - Rapid Onset - 50mg THC":
                for col in range(1, 18):
                    row[col].fill = green_fill
            elif third_cell == "muze - 7g" or third_cell == "muze - 1g (2x .5g) pre-rolls":
                for col in range(1, 18):
                    row[col].fill = purple_fill
            elif third_cell is not None:
                for col in range(1, 18):
                    row[col].fill = red_fill

    return sheet

# change the category separator font color to white, size to 13, font type and bold it
def update_cat_white(sheet):

    white_bold_font = Font(name='Calibri', size=13, bold=True, italic=False, color='FFFFFF')

    for row in sheet.iter_rows(min_row=8):
        if row[1].value is None or row[1].value == "":
                for cell in row:
                    cell.font = white_bold_font

def create_header(sheet):
    today = datetime.now()
    date = today.strftime('%m/%d/%Y')
    right_align = Alignment(horizontal='right')
    wg_cell = sheet['C2']
    wg_cell.value = 'WHOLESALE GOODS'
    wg_cell.alignment = right_align
    os_cell = sheet['C3']
    os_cell.value = 'ORDERING SHEET'
    os_cell.alignment = right_align
    date_cell = sheet['Q2']
    date_cell.value = date

    return sheet

def remove_zeros(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 0:
                cell.value = ' '

# available case formula - qty avail / case count
def available_case(sheet):
    for row in range(9, sheet.max_row + 1):
        n_value = sheet[f'N{row}'].value

        if isinstance(n_value, int):
            formula = f"=N{row}/M{row}"
            sheet[f'O{row}'] = formula

def convert_currency(sheet, column):

    currency_format='"$"#,##0.00'
    for row in range(9, sheet.max_row + 1):
        col_row = sheet[f'{column}{row}'].value

        if col_row:
            sheet[f'{column}{row}'].number_format = currency_format

# case price - price each * case count
def case_price(sheet):
    for row in range(9, sheet.max_row + 1):
        l_value = sheet[f'L{row}'].value
        # print(f'L Val: {l_value}')

        if l_value:
            formula = f"=L{row}*M{row}"
            sheet[f'P{row}'] = formula

# price/case * order quantity 
def item_total(sheet):
    for row in range(9, sheet.max_row + 1):
        p_value = sheet[f'P{row}'].value

        if p_value:
            formula = f'=IFERROR(IF(Q{row}="","$0.00",ROUND(Q{row}*P{row},2)),"0")'
            # =IFERROR(IF(Q11="","$0.00",ROUND(Q11*P11,2)),"0")
            sheet[f'R{row}'] = formula

# set column to word wrap and format text - yes, poorly named function
def word_wrap_column(ws, column):
    for cell in ws[column]:
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.font = Font(name='Calibri', size=12, bold=True, italic=False, color='000000')

def remove_border(sheet):
    no_border = Border(
    # left=Side(border_style=None),
    right=Side(border_style=None),
    top=Side(border_style=None),
    bottom=Side(border_style=None)
)
    cell = sheet['S7']
    cell.border = no_border

def update_color_in_column(sheet, column, color):
    fill_color = PatternFill(fill_type='solid', start_color=color, end_color=color)

    column_letter = column
    start_row = 9

    column_index = openpyxl.utils.column_index_from_string(column_letter)

    for row in sheet.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index):
        cell = row[0]
        if cell.value is not None:
            cell.fill = fill_color

# determines the row size of each category by using the category header row bg colors as start/stop points
# then drops the sum formula in the last row of the product category to be used in volume_pricing_ea_column
# returning the pd_cell_dictionary dictionary that assigns the total sum cell in column U
# to each product description. this dictionary will be passed to volume_pricing_ea_column for building the 
# volume pricing formula 
def insert_section_sums(sheet, last_row):

    pd_cell_dictionary = {}
    header_colors = {"00D60000", "0000B050", "007030A0",
                     "FFD60000", "FF00B050", "FF7030A0"}
    col_b = 2  # product description
    col_c = 3  # technically the strain column but leveraging for checking if header row
    col_q = 17  # order qty case - to build the sum formula going in U column
    col_u = 21  # to be hidden column with the sum total formula for each product category

    section_start = None

    for row in range(7, last_row + 1):
        cell = sheet.cell(row = row, column = col_c)
        pd_cell = sheet.cell(row = row - 1, column = col_b)
        fill = cell.fill

        # check for red-green-purple section row
        is_section_header = (
            fill and fill.fill_type == 'solid' and fill.start_color
            and fill.start_color.rgb in header_colors
        )

        if is_section_header:
            if section_start is not None:
                section_end = row - 1
                if section_end >= section_start:
                    sum_cell = sheet.cell(row=section_end, column=col_u)
                    sum_cell.value = f"=SUM({get_column_letter(col_q)}{section_start}:{get_column_letter(col_q)}{section_end})"
                    pd_cell_dictionary[pd_cell.value] = sum_cell.coordinate
            section_start = row + 1

    return pd_cell_dictionary

# builds volume pricing formula that's dropped in column T
# and updates the Total column (R) to look for volume pricing
def volume_pricing_ea_column(sheet, dict):
    vol_price_dict = cs.volume_pricing
    val_price_dict = cs.value_pricing

    for row in range(8, sheet.max_row + 1):
        prod_desc = sheet[f'B{row}'].value
        strain = sheet[f'C{row}'].value
        key_tuple = (prod_desc, strain)
        volume_price = sheet[f'L{row}']

        for key, val in vol_price_dict.items():
            if prod_desc == key and key_tuple not in val_price_dict:
                total_cell = dict.get(key)
                # vol_price_ea = f'=IF({total_cell}>=10, {val}, L{row}*M{row})'
                vol_price_ea = f'=IF({total_cell}>=10, {val}, T{row})'
                volume_price.value = vol_price_ea


def dupe_column(sheet, source_column, destination_column):

    max_row = sheet.max_row

    for row in range(8, max_row + 1):
        source_cell = sheet[f"{source_column}{row}"]
        destination_cell = sheet[f"{destination_column}{row}"]
        destination_cell.value = source_cell.value