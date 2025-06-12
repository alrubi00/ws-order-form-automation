import pandas as pd
import constants as cs
import df_functions as dfuns
import xlsx_functions as xfuns
import functions as funs
import acumatica as acu
import sharepoint as sp
import email_w_attach as ewa
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from datetime import datetime
from openpyxl import load_workbook
import os

####################################
## SET UP PATHS AND STARTING XLSX ##
####################################

script_dir = os.path.dirname(os.path.abspath(__file__))
tmp_path = os.path.abspath(os.path.join(script_dir, '..', 'tmp'))
data_path = os.path.abspath(os.path.join(script_dir, '..', 'data'))
img_path = os.path.abspath(os.path.join(script_dir, '..', 'img'))

ws_order_form_output = funs.join_dir_file(data_path, cs.ws_order_form_name)
create_logo_path = funs.join_dir_file(img_path, cs.logo_file_name)
create_tmp_xlsx_path = funs.join_dir_file(tmp_path, cs.tmp_xlsx_name)

img = Image(create_logo_path)
file_name = create_tmp_xlsx_path

wb = Workbook()
ws = wb.active
xfuns.format_white_bg(ws, 'A1:FF550')
ws.add_image(img, 'C2')
ws.title = cs.sheet_name
wb.save(file_name)

######################
## EXTRACTION START ##
######################

# login
# 3 calls to generate each report
# 3 subsequent calls to download the generated report in xlsx format
# logout

# RESOURCES DOWNLOADED 
# df_batch - WS050925 - FGA Wholesale With Batch Info No Group or Sort - CLAEBAvailableNoGroup
# df_qty -  WS040125 - FGA Wholesale With Valid Qty - FGAWSOF
# df_in_transit - WS040725 - In-Transit Inventory - INTRANSITINV

# create dataframe for - WS040725 - In-Transit Inventory - INTRANSITINV
df_in_transit = funs.login_generate_download_report_df('INTRANSITINV')

# create dataframe for WS050925 - FGA Wholesale With Batch Info No Group or Sort - CLAEBAvailableNoGroup
df_batch_grouped = funs.login_generate_download_report_df('CLAEBAvailableNoGroup')

# create dataframe for WS040125 - FGA Wholesale With Valid Qty - FGAWSOF
df_qty = funs.login_generate_download_report_df('FGAWSOF')

####################
## EXTRACTION END ##
####################

####################################
## DATAFRAME TRANSFORMATION START ##
####################################

# now start building main dataframe by merging the 3 dataframes from above extraction
main_df = dfuns.merge_dfs(df_qty, df_batch_grouped, df_in_transit)

# in case you couldn't tell by the function name :), removing dupe rows
main_df = dfuns.drop_dupe_rows(main_df)

# remove row if product has 0 for quantity
main_df = dfuns.remove_row_with_zero_qty(main_df, 'Qty Available for Sale')

# remove the strains not for general whoesale
main_df = dfuns.remove_row_with_val_in_col(main_df, 'Strain', cs.strain_no_sale_list)

# sort by Inventory ID and in specified order
sorted_df = dfuns.order_by_inventory_id(main_df)

# update harvest date's date format
sorted_df['Harvest Date'] = sorted_df['Harvest Date'].dt.strftime('%m/%d/%Y')

# some old harvest dates can sneak into the dataset - generally edible or extracts or muze - they aren't to be published  
sorted_df = dfuns.remove_old_dates(sorted_df, 'Harvest Date')

# remove harvest date from variety pack
sorted_df.loc[sorted_df['Inventory ID'].isin(['PR5-2.5', 'PR5-5']), 'Harvest Date'] = ''

# column header updates
cleaned_cols_df = sorted_df.rename(columns={'Strain': 'Strain/Flavor', 'THCA': 'THC-A', 'Qty Available for Sale': 'Qty. Available'})

## the next 4 steps add coluns to the df from hardcoded dictionaries
## this accomodates values that either aren't in acumatica or are really tough to get out acu
# add column with I/S/H value
ish_col_added_df = dfuns.add_col_with_vals_from_dict(cleaned_cols_df, 'Strain/Flavor', cs.ish_dict, 'I/S/H')

# add net weights/volumes column with value
net_col_added_df = dfuns.add_col_with_vals_from_dict(ish_col_added_df, 'Inventory ID', cs.net_weight_vol, 'Net Weights/Volumes')

# add servings column with value
serve_col_added_df = dfuns.add_col_with_vals_from_dict(net_col_added_df, 'Inventory ID', cs.servings, 'Servings')

# add price column with value
price_col_added_df = dfuns.add_col_with_vals_from_dict(serve_col_added_df, 'Inventory ID', cs.price_ea, 'Price/EA')

# add volume pricing
price_col_added_df = dfuns.add_col_with_vals_from_dict(price_col_added_df, 'Inventory ID', cs.volume_pricing_ad, ' ')

# this accomodates when the wholesale team wants a product to be on sale or have value pricing
price_col_added_df['Price/EA'] = price_col_added_df.apply(dfuns.value_pricing_update, axis=1)

# add case count column with value
case_count_col_added_df = dfuns.add_col_with_vals_from_dict(price_col_added_df, 'Inventory ID', cs.case_count, 'Case Count')

# convert qty to fit in case count
case_count_col_added_df = dfuns.qty_case_count_conv(case_count_col_added_df)

# the orginal qty available can be removed because it was converted above to a new column
case_count_col_added_df.pop('Qty. Available')

# rename the Conversion column back to Available
case_count_col_added_df = case_count_col_added_df.rename(columns={'Qty Conversion': 'Qty. Available'})

# then move I/S/H column to 3rd column
move_ish_col_added_df = dfuns.move_column(case_count_col_added_df, 'I/S/H', 3)

# then move Available column to 3rd column
move_qty_col_added_df = dfuns.move_column(move_ish_col_added_df, 'Qty. Available', 13)

# remove batch details from edibles
move_qty_col_added_df = dfuns.remove_batch_details(move_qty_col_added_df)

# update cfx gummy strain/flavor to include Sleep-Calm-Focus-Energy
move_qty_col_added_df = dfuns.update_cfx_gummies_description(move_qty_col_added_df)

# add cfx gummy cbd detail
move_qty_col_added_df = dfuns.add_value_to_col_based_on_other_col(move_qty_col_added_df, 'Total THC', cs.cfx_gum_cbds_map, 'Strain/Flavor')

# add tincture/topical thc/cbd values
move_qty_col_added_df = dfuns.add_value_to_col_based_on_other_col(move_qty_col_added_df, 'Total THC', cs.top_tinc_thc_cbd_map, 'Product Description')

# remove rows with nulls in Inventory ID and Qty. Available
final_df = move_qty_col_added_df[move_qty_col_added_df['Inventory ID'].notna()]
final_df = move_qty_col_added_df[move_qty_col_added_df['Qty. Available'].notna()]

# add columns not in df already (NET WEIGHTS/VOLUMES, SERVINGS, etc.)
add_columns_df = dfuns.add_columns(final_df)

# sort alphabetically by strain within each category
add_columns_df = add_columns_df.sort_values(by=['Inventory ID', 'Strain/Flavor'], ascending=[True, True])

# add product descriptor row to separate items (e.g. flower from pre-rolls from hitmakers and so on)
starting_row_cat_insert_df = dfuns.insert_start_row(add_columns_df, cs.cat_by_inventory_id)

# now we can remove the inventory id column - no longer needed
starting_row_cat_insert_df.pop('Inventory ID')

####################################
## DATAFRAME TRANSFORMATION END ##
####################################

# write the finished dataframe to the excel file
with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    starting_row_cat_insert_df.to_excel(writer, sheet_name=cs.sheet_name, startrow=6, startcol=1, index=False)


###############################
## XLSX TRANSFORMATION START ##
###############################

# start a new workbook for the remaining formatting
new_workbook = load_workbook(file_name)
sheet = new_workbook[cs.sheet_name]

# remaining format updates to the sheet
xfuns.grey_headers(sheet)
xfuns.delete_dupe_red_rows(sheet)
xfuns.update_cat_white(sheet)
xfuns.remove_zeros(sheet)
xfuns.convert_float_percentage(sheet)

# Part 1: merge the cannabinoid breakdown cell for Topical/Tinctures in column G into columns F and H (that are empty)
# so the Total THC column isn't significantly wider
get_top_coords = xfuns.get_product_coordinates(sheet, cs.cat_by_inventory_id['REMT-1:1-250'], cs.cat_by_inventory_id['HVG-6-SEEDPACK-AUTO'])
xfuns.merge_cbds_breakdown_cells(sheet, get_top_coords, cs.cat_by_inventory_id['REMT-1:1-250'], cs.cat_by_inventory_id['HVG-6-SEEDPACK-AUTO'])

# Part 2: merge the cannabinoid breakdown cell for CFX Gummies in column G into columns F and H (that are empty)
# so the Total THC column isn't significantly wider
get_cfxgum_coords = xfuns.get_product_coordinates(sheet, cs.cat_by_inventory_id['GUM100-CFX-CALM'], cs.cat_by_inventory_id['GUM100-RAPID'])
xfuns.merge_cbds_breakdown_cells(sheet, get_cfxgum_coords, cs.cat_by_inventory_id['GUM100-CFX-CALM'], cs.cat_by_inventory_id['GUM100-RAPID'])

# link the strains in FLOWER - Jar 3.5g (+) to their cultivar page on hv.org
# get_flwr35_coords = xfuns.get_product_coordinates(sheet, cs.cat_by_inventory_id['FLWR-3.5-PLUS'], cs.cat_by_inventory_id['PR1'])
# xfuns.link_strain_to_cultivar(sheet, get_flwr35_coords, cs.strain_to_cult_page, cs.cat_by_inventory_id['FLWR-3.5-PLUS'], cs.cat_by_inventory_id['PR1'])

# link the seed strains in both seed pack sections to their genetics page on hv.org
get_seed_coords = xfuns.get_product_coordinates(sheet, cs.cat_by_inventory_id['HVG-6-SEEDPACK-AUTO'], cs.cat_by_inventory_id['HM-DSP-LVO-1G'])
xfuns.link_strain_to_genetics(sheet, get_seed_coords, cs.strain_to_gen_page, cs.cat_by_inventory_id['HVG-6-SEEDPACK-AUTO'], cs.cat_by_inventory_id['HM-DSP-LVO-1G'])

# xfuns.link_strain_to_cultivar(sheet, cs.strain_to_page)
xfuns.adjust_column_width(sheet)
xfuns.center_align_columns(sheet)
xfuns.update_value_pricing_bg(sheet)
xfuns.available_case(sheet)
xfuns.convert_currency(sheet, 'L')
xfuns.case_price(sheet)
xfuns.item_total(sheet)
xfuns.convert_currency(sheet, 'P')
last_total_row = xfuns.get_max_total_row(sheet)
xfuns.add_borders(sheet, last_total_row)
xfuns.grey_out_cells(sheet, last_total_row)
xfuns.add_separator_row(sheet)
xfuns.add_total_sum(sheet, last_total_row)
pd_sum_cell_dictionary = xfuns.insert_section_sums(sheet, last_total_row)
xfuns.dupe_column(sheet, 'L', 'T')
xfuns.volume_pricing_ea_column(sheet, pd_sum_cell_dictionary)
xfuns.merge_cells_in_column(sheet, 'B', 9)
xfuns.convert_currency(sheet, 'R')
xfuns.update_color_in_column(sheet, 'S', 'FCE4D6')
xfuns.merge_cells_in_column(sheet, 'S', 9)
# widening the Total column so larger totals doesn't
# get converted 'visually' to #### because the column width is too narrow
sheet.column_dimensions['R'].width = 15
# widening the Volume Pricing column to fit Volume Pricing messaging
sheet.column_dimensions['S'].width = 27
# setting coa columns to be uniform
sheet.column_dimensions['E'].width = 16
sheet.column_dimensions['F'].width = 16
sheet.column_dimensions['G'].width = 16
sheet.column_dimensions['H'].width = 16
xfuns.word_wrap_column(sheet, 'S')
xfuns.remove_border(sheet)
sheet.column_dimensions['B'].hidden = True
sheet.column_dimensions['T'].hidden = True
sheet.column_dimensions['U'].hidden = True
xfuns.create_header(sheet)
sheet.sheet_view.zoomScale = 75

# freeze first 7 rows so you don't loose column headers
sheet.freeze_panes = 'A8'

# new_workbook.save(f'../data/Wholesale_Order_Form_{date_for_file}.xlsx')
new_workbook.save(ws_order_form_output)

#############################
## XLSX TRANSFORMATION END ##
#############################

#######################################
## EMAIL FORM/ADD FORM TO SHAREPOINT ##
#######################################

# add form to sharepoint
link_to_file_on_sp = sp.add_form_to_sharepoint(ws_order_form_output)

# email the output form
# gwa.send_email(ws_order_form_output)
ewa.email_form_w_link(ws_order_form_output, link_to_file_on_sp)

##############
## CLEAN UP ##
##############

# clean up tmp files
funs.delete_files_from_directory(tmp_path)

# clean out older files
funs.delete_old_files(data_path)